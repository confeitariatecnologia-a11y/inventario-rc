import re
import openpyxl
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
EXCEL_PATH = ROOT.parent / "Imobilizado Richesse (1).xlsx"
if not EXCEL_PATH.exists():
    EXCEL_PATH = ROOT / "temp_imobilizado.xlsx"

OUTPUT_SQL = ROOT / "supabase" / "import_imobilizado_completo.sql"

wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb["Analítico"]

LOCATION_MAP = {
    "OESTE": "Oeste",
    "MARISTA": "Marista",
    "PRIME": "Prime",
    "FLAMBOYANT": "Flamboyant",
    "BRASILIA": "Brasília",
    "BRASÍLIA": "Brasília",
    "TO GO": "Togo",
    "GYN SHOPPING": "Goiânia Shopping",
    "GOIÂNIA SHOPPING": "Goiânia Shopping",
    "GOIANIA SHOPPING": "Goiânia Shopping",
}

CATEGORY_MAP = {
    "MÓVEIS E UTENSÍLIOS": "Móveis e Utensílios",
    "MOVEIS E UTENSILIOS": "Móveis e Utensílios",
    "MÁQUINAS E EQUIPAMENTOS": "Máquinas e Equipamentos",
    "MAQUINAS E EQUIPAMENTOS": "Máquinas e Equipamentos",
    "COMPUTADORES E PERIFÉRICOS": "Computadores e Periféricos",
    "COMPUTADORES E PERIFERICOS": "Computadores e Periféricos",
    "VEÍCULOS": "Veículos",
    "VEICULOS": "Veículos",
}

def escape_sql(val):
    if val is None:
        return "NULL"
    s = str(val).strip().replace("'", "''")
    return f"'{s}'"

def escape_num(val):
    if val is None or not isinstance(val, (int, float)):
        return "NULL"
    return str(round(float(val), 2))

rows = []
for r in range(2, ws.max_row + 1):
    item_id = ws.cell(r, 1).value
    grupo = ws.cell(r, 2).value
    plaqueta = ws.cell(r, 3).value
    descricao = ws.cell(r, 4).value
    loja = ws.cell(r, 5).value
    valor = ws.cell(r, 7).value
    obs = ws.cell(r, 9).value

    if not descricao and not grupo and not loja:
        continue

    cat_name = CATEGORY_MAP.get(str(grupo).strip().upper(), str(grupo).strip())
    loc_name = LOCATION_MAP.get(str(loja).strip().upper(), str(loja).strip())
    
    pat_code = f"PAT-{len(rows) + 1:04d}"
    plaq_str = str(plaqueta).strip() if plaqueta is not None else ""
    
    notes_parts = []
    if plaq_str:
        notes_parts.append(f"Plaqueta Original: {plaq_str}")
    if grupo:
        notes_parts.append(f"Grupo: {grupo}")
    if obs:
        notes_parts.append(f"Obs: {obs}")
    
    notes = " | ".join(notes_parts) if notes_parts else None

    rows.append({
        "asset_code": pat_code,
        "name": str(descricao).strip(),
        "serial_number": plaq_str,
        "category_name": cat_name,
        "location_name": loc_name,
        "acquisition_value": valor,
        "notes": notes,
    })

print(f"Total de ativos processados com sucesso: {len(rows)}")

sql_lines = [
    "-- ==============================================================================",
    f"-- IMPORTAÇÃO COMPLETA: {len(rows)} ATIVOS DE 'Imobilizado Richesse (1).xlsx'",
    "-- Execute no SQL Editor do Supabase (https://supabase.com/dashboard)",
    "-- ==============================================================================",
    "",
    "BEGIN;",
    "",
    "-- 1. Garantir Localizações (Unidades)",
    "INSERT INTO locations (name, type) VALUES",
    "  ('Oeste', 'loja'),",
    "  ('Marista', 'loja'),",
    "  ('Prime', 'loja'),",
    "  ('Flamboyant', 'loja'),",
    "  ('Brasília', 'loja'),",
    "  ('Togo', 'loja'),",
    "  ('Goiânia Shopping', 'loja')",
    "ON CONFLICT DO NOTHING;",
    "",
    "-- 2. Garantir Categorias",
    "INSERT INTO categories (name, icon, color) VALUES",
    "  ('Móveis e Utensílios', 'Armchair', '#64748b'),",
    "  ('Máquinas e Equipamentos', 'Cog', '#f59e0b'),",
    "  ('Computadores e Periféricos', 'Monitor', '#2563eb'),",
    "  ('Veículos', 'Truck', '#ef4444')",
    "ON CONFLICT DO NOTHING;",
    "",
    "-- 3. Tabela Temporária para Carga Rápida",
    "CREATE TEMP TABLE tmp_import (",
    "  asset_code text,",
    "  name text,",
    "  serial_number text,",
    "  category_name text,",
    "  location_name text,",
    "  acquisition_value numeric,",
    "  notes text",
    ") ON COMMIT DROP;",
    "",
    "INSERT INTO tmp_import VALUES"
]

val_tuples = []
for row in rows:
    t = f"  ({escape_sql(row['asset_code'])}, {escape_sql(row['name'])}, {escape_sql(row['serial_number'])}, {escape_sql(row['category_name'])}, {escape_sql(row['location_name'])}, {escape_num(row['acquisition_value'])}, {escape_sql(row['notes'])})"
    val_tuples.append(t)

sql_lines.append(",\n".join(val_tuples) + ";")

sql_lines.extend([
    "",
    "-- 4. Upsert na tabela assets vinculando IDs de Categoria e Unidade",
    "INSERT INTO assets (",
    "  asset_code, name, serial_number, category_id, location_id, status, acquisition_value, notes, qr_code, updated_at",
    ")",
    "SELECT",
    "  t.asset_code,",
    "  t.name,",
    "  NULLIF(t.serial_number, ''),",
    "  c.id,",
    "  l.id,",
    "  'operacional',",
    "  t.acquisition_value,",
    "  t.notes,",
    "  t.asset_code,",
    "  NOW()",
    "FROM tmp_import t",
    "LEFT JOIN categories c ON c.name = t.category_name",
    "LEFT JOIN locations l ON l.name = t.location_name",
    "ON CONFLICT (asset_code) DO UPDATE SET",
    "  name = EXCLUDED.name,",
    "  serial_number = EXCLUDED.serial_number,",
    "  category_id = EXCLUDED.category_id,",
    "  location_id = EXCLUDED.location_id,",
    "  acquisition_value = EXCLUDED.acquisition_value,",
    "  notes = EXCLUDED.notes,",
    "  qr_code = EXCLUDED.qr_code,",
    "  updated_at = NOW();",
    "",
    "COMMIT;",
    "",
    "-- Conferência:",
    "-- SELECT count(*) FROM assets;"
])

OUTPUT_SQL.write_text("\n".join(sql_lines), encoding="utf-8")
print(f"Arquivo gerado: {OUTPUT_SQL} ({OUTPUT_SQL.stat().st_size / 1024:.1f} KB)")
